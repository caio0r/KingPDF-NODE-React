from fastapi import APIRouter, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
import shutil
import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_LEFT
from core.utils import cleanup_file

router = APIRouter()

UPLOAD_DIR = "uploads"
OUTPUT_DIR = "outputs"

@router.post("/excel-to-pdf")
async def excel_to_pdf(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload an Excel file (.xlsx or .xls).")

    # Note: .xls files (old format) are not fully supported, primarily .xlsx
    if file.filename.endswith(".xls"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported. Please convert your .xls file to .xlsx first.")

    input_path = os.path.join(UPLOAD_DIR, f"excel_in_{file.filename}")
    output_filename = f"{os.path.splitext(file.filename)[0]}.pdf"
    output_path = os.path.join(OUTPUT_DIR, output_filename)

    try:
        # Save uploaded file
        with open(input_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Load Excel workbook
        wb = load_workbook(input_path, data_only=True)
        
        # Create PDF with landscape orientation (better for spreadsheets)
        pdf = SimpleDocTemplate(
            output_path, 
            pagesize=landscape(A4),
            rightMargin=0.3*inch, 
            leftMargin=0.3*inch,
            topMargin=0.5*inch, 
            bottomMargin=0.5*inch
        )
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        
        # Create custom style for table cells
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            alignment=TA_LEFT,
            wordWrap='CJK',
        )
        
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            alignment=TA_LEFT,
            textColor=colors.whitesmoke,
            fontName='Helvetica-Bold',
        )
        
        sheet_count = 0
        
        # Process each worksheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_count += 1
            
            # Add sheet name as title
            sheet_title = Paragraph(f"Planilha: {sheet_name}", title_style)
            elements.append(sheet_title)
            elements.append(Spacer(1, 12))
            
            # Get all data from sheet
            data = []
            max_col = 0
            
            # First, collect all data
            for row in sheet.iter_rows(values_only=True):
                # Convert None to empty string and all values to strings
                row_data = [str(cell) if cell is not None else "" for cell in row]
                
                # Track maximum columns
                if len(row_data) > max_col:
                    max_col = len(row_data)
                
                # Only add non-empty rows
                if any(cell for cell in row_data):
                    data.append(row_data)
            
            # Normalize all rows to have same number of columns
            for row in data:
                while len(row) < max_col:
                    row.append("")
            
            if data and max_col > 0:
                try:
                    # Calculate column widths dynamically
                    available_width = 10.5 * inch  # Total available width
                    
                    # Analyze content to determine optimal column widths
                    col_widths = []
                    for col_idx in range(max_col):
                        max_length = 0
                        for row in data[:20]:  # Sample first 20 rows
                            if col_idx < len(row):
                                cell_length = len(row[col_idx])
                                if cell_length > max_length:
                                    max_length = cell_length
                        
                        # Base width on content length, with min and max limits
                        width = min(max(0.8 * inch, max_length * 0.05 * inch), 3 * inch)
                        col_widths.append(width)
                    
                    # Normalize widths to fit available space
                    total_width = sum(col_widths)
                    if total_width > available_width:
                        scale = available_width / total_width
                        col_widths = [w * scale for w in col_widths]
                    
                    # Convert data to Paragraphs for better text wrapping
                    formatted_data = []
                    for row_idx, row in enumerate(data):
                        formatted_row = []
                        for cell_text in row:
                            # Use header style for first row, cell style for others
                            style = header_style if row_idx == 0 else cell_style
                            
                            # Clean text and create Paragraph
                            clean_text = str(cell_text).strip()
                            if clean_text:
                                # Replace line breaks with <br/> for reportlab
                                clean_text = clean_text.replace('\n', '<br/>')
                                para = Paragraph(clean_text, style)
                            else:
                                para = Paragraph("", style)
                            formatted_row.append(para)
                        formatted_data.append(formatted_row)
                    
                    # Create table
                    t = Table(formatted_data, colWidths=col_widths)
                    
                    # Style the table
                    table_style = [
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 9),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                        ('TOPPADDING', (0, 0), (-1, 0), 8),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('LEFTPADDING', (0, 0), (-1, -1), 4),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                        ('TOPPADDING', (0, 1), (-1, -1), 4),
                        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                    ]
                    
                    t.setStyle(TableStyle(table_style))
                    elements.append(t)
                    
                except Exception as e:
                    print(f"Error creating table for sheet {sheet_name}: {e}")
                    import traceback
                    traceback.print_exc()
                    # Add error message
                    error_para = Paragraph(f"Erro ao processar planilha: {str(e)}", styles['Normal'])
                    elements.append(error_para)
            else:
                # Add empty sheet message
                empty_para = Paragraph("Planilha vazia", styles['Normal'])
                elements.append(empty_para)
            
            # Add page break between sheets (except for the last one)
            if sheet_count < len(wb.sheetnames):
                elements.append(PageBreak())
        
        # Build PDF
        if elements:
            pdf.build(elements)
            print(f"Successfully converted {sheet_count} sheets to PDF")
        else:
            # Create empty message
            elements.append(Paragraph("O arquivo Excel não contém dados.", styles['Normal']))
            pdf.build(elements)

        if not os.path.exists(output_path):
            raise HTTPException(status_code=500, detail="Conversion failed: Output file not created.")

        background_tasks.add_task(cleanup_file, output_path)

        return FileResponse(
            output_path, 
            media_type="application/pdf", 
            filename=output_filename
        )

    except Exception as e:
        print(f"Conversion error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")
    
    finally:
        cleanup_file(input_path)
